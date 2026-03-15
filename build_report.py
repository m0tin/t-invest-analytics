"""
T-Invest Analytics Report Builder v2
Мощный инвестиционный Excel-отчёт с аналитикой Bloomberg-уровня.
Данные загружаются через T-Invest REST API (read-only).
Период анализа: с 01.01.2024.
"""
import os
import sys
import time
from bisect import bisect_right
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from collections import defaultdict, deque

import requests
from dotenv import load_dotenv
from scipy.optimize import brentq

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import (
    PieChart, BarChart, LineChart, AreaChart, Reference,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule

load_dotenv()

TOKEN = os.getenv("TINKOFF_TOKEN")
BASE_URL = "https://invest-public-api.tinkoff.ru/rest"
HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Content-Type": "application/json",
}

ANALYSIS_START = datetime(2024, 1, 1, tzinfo=timezone.utc)

# ── Operation type groups (used instead of raw strings) ──
BUY_SELL_TYPES = {"OPERATION_TYPE_BUY", "OPERATION_TYPE_SELL"}
DEPOSIT_TYPES = {"OPERATION_TYPE_INPUT", "OPERATION_TYPE_INP_MULTI"}
WITHDRAWAL_TYPES = {"OPERATION_TYPE_OUTPUT", "OPERATION_TYPE_OUT_MULTI"}
COMMISSION_TYPES = {"OPERATION_TYPE_BROKER_FEE", "OPERATION_TYPE_SERVICE_FEE"}

# ── Styles (строгий финансовый стиль) ──────────────────────────────────
HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL_GREEN = HEADER_FILL
HEADER_FILL_ORANGE = HEADER_FILL
HEADER_FILL_RED = HEADER_FILL
HEADER_FILL_PURPLE = HEADER_FILL
LIGHT_GREY_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
KPI_FONT = Font(bold=True, size=13)
KPI_LABEL_FONT = Font(color="808080", size=8)
TITLE_FONT = Font(bold=True, size=12)
SUBTITLE_FONT = Font(bold=True, size=10)
BOTTOM_BORDER = Border(bottom=Side(style="thin", color="BFBFBF"))
THIN_BORDER = Border(
    bottom=Side(style="hair", color="D9D9D9"),
)
COLOR_POS = "2E7D32"  # dark green for positive
COLOR_NEG = "C62828"  # dark red for negative
COLOR_LINE = "37474F"  # dark grey for charts
COLOR_AREA = "B0BEC5"  # light grey for area fills
RUB_FMT = '#,##0.00 "₽"'
RUB_FMT_INT = '#,##0 "₽"'
PCT_FMT = '0.00%'
DATE_FMT = 'DD.MM.YYYY'
DATETIME_FMT = 'DD.MM.YYYY HH:MM'

OPERATION_TYPES = {
    "OPERATION_TYPE_BUY": "Покупка",
    "OPERATION_TYPE_SELL": "Продажа",
    "OPERATION_TYPE_COUPON": "Купоны",
    "OPERATION_TYPE_DIVIDEND": "Дивиденды",
    "OPERATION_TYPE_DIVIDEND_TAX": "Налог на дивиденды",
    "OPERATION_TYPE_BROKER_FEE": "Комиссия брокера",
    "OPERATION_TYPE_TAX": "Налог",
    "OPERATION_TYPE_BOND_REPAYMENT": "Погашение облигации",
    "OPERATION_TYPE_BOND_REPAYMENT_FULL": "Погашение облигации",
    "OPERATION_TYPE_INPUT": "Пополнение",
    "OPERATION_TYPE_OUTPUT": "Вывод",
    "OPERATION_TYPE_OVERNIGHT": "Овернайт",
    "OPERATION_TYPE_ACCRUING_VARMARGIN": "Вариационная маржа",
    "OPERATION_TYPE_WRITING_OFF_VARMARGIN": "Списание вариационной маржи",
    "OPERATION_TYPE_TAX_CORRECTION": "Корректировка налога",
    "OPERATION_TYPE_BENEFIT_TAX": "Налог на доход",
    "OPERATION_TYPE_SERVICE_FEE": "Сервисная комиссия",
    "OPERATION_TYPE_INP_MULTI": "Пополнение (мульти)",
    "OPERATION_TYPE_OUT_MULTI": "Вывод (мульти)",
    "OPERATION_TYPE_BOND_TAX": "Налог на купон",
}

OPERATION_GROUPS = {
    "OPERATION_TYPE_BUY": "Торговля",
    "OPERATION_TYPE_SELL": "Торговля",
    "OPERATION_TYPE_COUPON": "Доход",
    "OPERATION_TYPE_DIVIDEND": "Доход",
    "OPERATION_TYPE_DIVIDEND_TAX": "Налоги",
    "OPERATION_TYPE_BROKER_FEE": "Комиссии",
    "OPERATION_TYPE_TAX": "Налоги",
    "OPERATION_TYPE_BOND_REPAYMENT": "Доход",
    "OPERATION_TYPE_BOND_REPAYMENT_FULL": "Доход",
    "OPERATION_TYPE_INPUT": "Денежные потоки",
    "OPERATION_TYPE_OUTPUT": "Денежные потоки",
    "OPERATION_TYPE_OVERNIGHT": "Прочее",
    "OPERATION_TYPE_ACCRUING_VARMARGIN": "Торговля",
    "OPERATION_TYPE_WRITING_OFF_VARMARGIN": "Торговля",
    "OPERATION_TYPE_TAX_CORRECTION": "Налоги",
    "OPERATION_TYPE_BENEFIT_TAX": "Налоги",
    "OPERATION_TYPE_SERVICE_FEE": "Комиссии",
    "OPERATION_TYPE_INP_MULTI": "Денежные потоки",
    "OPERATION_TYPE_OUT_MULTI": "Денежные потоки",
    "OPERATION_TYPE_BOND_TAX": "Налоги",
}

INCOME_TYPES = {
    "OPERATION_TYPE_COUPON", "OPERATION_TYPE_DIVIDEND",
    "OPERATION_TYPE_BOND_REPAYMENT", "OPERATION_TYPE_BOND_REPAYMENT_FULL",
}

TAX_TYPES = {
    "OPERATION_TYPE_TAX", "OPERATION_TYPE_DIVIDEND_TAX",
    "OPERATION_TYPE_BOND_TAX", "OPERATION_TYPE_TAX_CORRECTION",
    "OPERATION_TYPE_BENEFIT_TAX",
}

INSTRUMENT_TYPE_MAP = {
    "share": "Акции",
    "bond": "Облигации",
    "etf": "Фонды",
    "currency": "Валюта",
    "futures": "Фьючерсы",
    "option": "Опционы",
    "sp": "Структурные продукты",
}

ACCOUNT_TYPE_MAP = {
    "ACCOUNT_TYPE_TINKOFF": "Брокерский",
    "ACCOUNT_TYPE_TINKOFF_IIS": "ИИС",
}


# ══════════════════════════════════════════════════════════════════════════
# API LAYER
# ══════════════════════════════════════════════════════════════════════════

SESSION = requests.Session()
SESSION.headers.update(HEADERS)


def api_call(service, method, body=None):
    url = f"{BASE_URL}/tinkoff.public.invest.api.contract.v1.{service}/{method}"
    resp = SESSION.post(url, json=body or {})
    if resp.status_code != 200:
        return None
    return resp.json()


def money_value(m):
    """Parse T-Invest MoneyValue/Quotation {units, nano} structure."""
    if not m:
        return 0.0
    return int(m.get("units", 0)) + int(m.get("nano", 0)) / 1_000_000_000


quotation_value = money_value


def parse_ts(ts_str):
    if not ts_str:
        return None
    try:
        dt = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
        return dt.replace(tzinfo=None)
    except Exception:
        return None


def get_accounts():
    data = api_call("UsersService", "GetAccounts")
    if not data:
        return []
    return [a for a in data.get("accounts", []) if a.get("status") == "ACCOUNT_STATUS_OPEN"]


def get_portfolio(account_id):
    return api_call("OperationsService", "GetPortfolio", {"accountId": account_id, "currency": "RUB"})


def get_operations(account_id, from_date, to_date):
    all_ops = []
    cursor = ""
    while True:
        body = {
            "accountId": account_id,
            "from": from_date.isoformat(),
            "to": to_date.isoformat(),
            "state": "OPERATION_STATE_EXECUTED",
            "limit": 1000,
            "operationTypes": [],
            "withoutCommissions": False,
            "withoutTrades": False,
            "withoutOvernights": True,
        }
        if cursor:
            body["cursor"] = cursor
        data = api_call("OperationsService", "GetOperationsByCursor", body)
        if not data:
            break
        all_ops.extend(data.get("items", []))
        if not data.get("hasNext", False):
            break
        cursor = data.get("nextCursor", "")
        if not cursor:
            break
    return all_ops


def get_instrument_by_uid(uid):
    data = api_call("InstrumentsService", "GetInstrumentBy", {
        "idType": "INSTRUMENT_ID_TYPE_UID",
        "id": uid,
    })
    return data.get("instrument") if data else None


def get_candles(figi, from_date, to_date, interval="CANDLE_INTERVAL_DAY"):
    """Fetch candles, chunking by year for daily interval."""
    all_candles = []
    current = from_date
    while current < to_date:
        chunk_end = min(current + timedelta(days=364), to_date)
        data = api_call("MarketDataService", "GetCandles", {
            "figi": figi,
            "from": current.isoformat(),
            "to": chunk_end.isoformat(),
            "interval": interval,
        })
        if data:
            all_candles.extend(data.get("candles", []))
        time.sleep(0.3)
        current = chunk_end + timedelta(days=1)
    return all_candles


# ══════════════════════════════════════════════════════════════════════════
# CALCULATIONS
# ══════════════════════════════════════════════════════════════════════════

def xirr(cashflows):
    """Calculate XIRR from list of (date, amount) tuples.
    Returns annualized return rate or None if cannot compute."""
    if len(cashflows) < 2:
        return None
    # Ensure we have both positive and negative cashflows
    has_pos = any(cf[1] > 0 for cf in cashflows)
    has_neg = any(cf[1] < 0 for cf in cashflows)
    if not (has_pos and has_neg):
        return None

    d0 = min(cf[0] for cf in cashflows)

    def npv(rate):
        return sum(cf / ((1 + rate) ** ((d - d0).days / 365.25))
                   for d, cf in cashflows)

    try:
        return brentq(npv, -0.99, 10.0, maxiter=1000)
    except (ValueError, RuntimeError):
        try:
            return brentq(npv, -0.5, 5.0, maxiter=1000)
        except (ValueError, RuntimeError):
            return None


def calculate_realized_pnl(operations):
    """FIFO-based realized P&L calculation. Returns list of closed lot dicts."""
    # Group buy/sell operations by ticker
    trades_by_ticker = defaultdict(list)
    for op in operations:
        if op["type"] in BUY_SELL_TYPES:
            trades_by_ticker[op["ticker"]].append(op)

    closed_lots = []
    for ticker, trades in trades_by_ticker.items():
        trades_sorted = sorted(trades, key=lambda x: x["date"] or datetime.min)
        fifo_queue = deque()  # (date, qty, price_per_unit, account)

        for t in trades_sorted:
            qty = abs(t.get("quantity", 0))
            if qty == 0:
                continue
            price = abs(t.get("price", 0))
            if price == 0 and qty > 0:
                price = abs(t.get("payment", 0)) / qty if qty else 0

            if t["type"] == "OPERATION_TYPE_BUY":
                fifo_queue.append({
                    "date": t["date"],
                    "qty": qty,
                    "price": price,
                    "account": t.get("account_name", ""),
                    "name": t.get("name", ""),
                })
            elif t["type"] == "OPERATION_TYPE_SELL" and fifo_queue:
                remaining = qty
                sell_price = price
                sell_date = t["date"]

                while remaining > 0 and fifo_queue:
                    lot = fifo_queue[0]
                    matched = min(remaining, lot["qty"])

                    pnl = matched * (sell_price - lot["price"])
                    days_held = (sell_date - lot["date"]).days if sell_date and lot["date"] else 0
                    annual_return = 0
                    if lot["price"] > 0 and days_held > 0:
                        annual_return = ((sell_price / lot["price"]) ** (365.0 / days_held) - 1)

                    closed_lots.append({
                        "ticker": ticker,
                        "name": lot["name"],
                        "account": lot["account"],
                        "buy_date": lot["date"],
                        "sell_date": sell_date,
                        "quantity": matched,
                        "buy_price": lot["price"],
                        "sell_price": sell_price,
                        "pnl": pnl,
                        "pnl_pct": (sell_price / lot["price"] - 1) if lot["price"] else 0,
                        "days_held": days_held,
                        "annual_return": annual_return,
                    })

                    lot["qty"] -= matched
                    remaining -= matched
                    if lot["qty"] <= 0:
                        fifo_queue.popleft()

    return closed_lots


def build_monthly_cashflows(operations):
    """Build monthly aggregated cash flow table."""
    months = defaultdict(lambda: {
        "deposits": 0, "withdrawals": 0, "dividends": 0,
        "coupons": 0, "commissions": 0, "taxes": 0, "trades": 0,
    })

    for op in operations:
        d = op.get("date")
        if not d:
            continue
        month_key = d.strftime("%Y-%m")
        payment = op.get("payment", 0)
        op_type = op.get("type", "")

        if op_type in DEPOSIT_TYPES:
            months[month_key]["deposits"] += payment
        elif op_type in WITHDRAWAL_TYPES:
            months[month_key]["withdrawals"] += payment
        elif op_type == "OPERATION_TYPE_DIVIDEND":
            months[month_key]["dividends"] += payment
        elif op_type == "OPERATION_TYPE_COUPON":
            months[month_key]["coupons"] += payment
        elif op_type in TAX_TYPES:
            months[month_key]["taxes"] += payment
        elif op_type in COMMISSION_TYPES:
            months[month_key]["commissions"] += payment
        elif op_type in BUY_SELL_TYPES:
            months[month_key]["trades"] += payment

    return dict(sorted(months.items()))


def aggregate_taxes(operations):
    """Aggregate taxes by year and type."""
    taxes = defaultdict(lambda: defaultdict(float))
    for op in operations:
        if op.get("type") not in TAX_TYPES:
            continue
        d = op.get("date")
        if not d:
            continue
        year = d.year
        tax_type = OPERATION_TYPES.get(op["type"], op["type"])
        taxes[year][tax_type] += abs(op.get("payment", 0))
    return dict(taxes)


def aggregate_dividends_by_ticker(operations):
    """Sum dividends received per ticker."""
    divs = defaultdict(float)
    for op in operations:
        if op.get("type") in INCOME_TYPES:
            divs[op.get("ticker", "")] += op.get("payment", 0)
    return dict(divs)


def trailing_12m_income(operations):
    """Calculate total income (dividends + coupons) for last 12 months."""
    t12m_start = datetime.now().replace(tzinfo=None) - timedelta(days=365)
    return sum(
        op.get("payment", 0) for op in operations
        if op.get("type") in INCOME_TYPES and op.get("date") and op["date"] >= t12m_start
    )


def calculate_monthly_returns(portfolio_history):
    """Calculate monthly returns adjusted for cash flows. Returns dict {(year, month): return}."""
    monthly_returns = {}
    for i in range(1, len(portfolio_history)):
        prev = portfolio_history[i - 1]
        curr = portfolio_history[i]
        if prev["value"] > 0:
            net_flow = (curr["cum_deposits"] - prev["cum_deposits"]
                        - (curr["cum_withdrawals"] - prev["cum_withdrawals"]))
            ret = (curr["value"] - prev["value"] - net_flow) / prev["value"]
            d = curr["date"]
            monthly_returns[(d.year, d.month)] = ret
    return monthly_returns


def calculate_drawdowns(portfolio_history):
    """Calculate drawdown from peak for each snapshot. Returns list of floats."""
    peak = 0
    drawdowns = []
    for h in portfolio_history:
        if h["value"] > peak:
            peak = h["value"]
        drawdowns.append((h["value"] / peak - 1) if peak > 0 else 0)
    return drawdowns


def reconstruct_portfolio_history(operations, candles_cache):
    """Reconstruct monthly portfolio snapshots from operations + candle prices.
    Returns list of dicts: {date, value, cum_deposits, cum_withdrawals}."""
    # Build position ledger over time
    ops_sorted = sorted(
        [op for op in operations if op.get("date")],
        key=lambda x: x["date"]
    )
    if not ops_sorted:
        return []

    positions = defaultdict(float)  # ticker -> qty
    cum_deposits = 0.0
    cum_withdrawals = 0.0

    # Find all month-end dates from start to now
    start = ANALYSIS_START.replace(tzinfo=None)
    end = datetime.now().replace(tzinfo=None)
    month_ends = []
    d = datetime(start.year, start.month, 1)
    while d <= end:
        next_month = d.month % 12 + 1
        next_year = d.year + (1 if d.month == 12 else 0)
        month_end = datetime(next_year, next_month, 1) - timedelta(days=1)
        if month_end <= end:
            month_ends.append(month_end)
        d = datetime(next_year, next_month, 1)

    # Add today as the last snapshot
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
    if not month_ends or month_ends[-1] != today:
        month_ends.append(today)

    history = []
    op_idx = 0

    for me in month_ends:
        # Apply all operations up to this month-end
        while op_idx < len(ops_sorted):
            op = ops_sorted[op_idx]
            if op["date"] > me:
                break
            op_type = op.get("type", "")
            ticker = op.get("ticker", "")
            raw_qty = op.get("quantity", 0)
            qty = abs(int(raw_qty)) if raw_qty and raw_qty != "" else 0

            if op_type == "OPERATION_TYPE_BUY" and ticker and qty:
                positions[ticker] += qty
            elif op_type == "OPERATION_TYPE_SELL" and ticker and qty:
                positions[ticker] -= qty
            elif op_type in DEPOSIT_TYPES:
                cum_deposits += op.get("payment", 0)
            elif op_type in WITHDRAWAL_TYPES:
                cum_withdrawals += abs(op.get("payment", 0))

            op_idx += 1

        # Calculate portfolio value using bisect for O(log N) price lookup
        total_value = 0.0
        for ticker, qty in positions.items():
            if qty <= 0:
                continue
            candles = candles_cache.get(ticker)
            if not candles:
                continue
            dates = candles["dates"]
            prices = candles["prices"]
            idx = bisect_right(dates, me) - 1
            price = prices[idx] if idx >= 0 else 0
            total_value += qty * price

        history.append({
            "date": me,
            "value": total_value,
            "cum_deposits": cum_deposits,
            "cum_withdrawals": cum_withdrawals,
        })

    return history


# ══════════════════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ══════════════════════════════════════════════════════════════════════════

def add_pnl_color_scale(ws, cell_range):
    """Subtle red-white-green color scale centered at 0."""
    ws.conditional_formatting.add(cell_range,
        ColorScaleRule(
            start_type='min', start_color='FFCDD2',
            mid_type='num', mid_value=0, mid_color='FFFFFF',
            end_type='max', end_color='C8E6C9'))


def add_portfolio_chart(ws, portfolio_history, data_start_row, chart_anchor, title="Стоимость портфеля"):
    """Write portfolio history data rows and create compact AreaChart."""
    ws.cell(row=data_start_row, column=1, value="Дата")
    ws.cell(row=data_start_row, column=2, value="Стоимость")
    ws.cell(row=data_start_row, column=3, value="Внесено")

    for i, h in enumerate(portfolio_history):
        r = data_start_row + 1 + i
        ws.cell(row=r, column=1, value=h["date"]).number_format = DATE_FMT
        ws.cell(row=r, column=2, value=h["value"]).number_format = RUB_FMT_INT
        ws.cell(row=r, column=3, value=h["cum_deposits"]).number_format = RUB_FMT_INT

    last_row = data_start_row + len(portfolio_history)

    chart = AreaChart()
    chart.title = title
    chart.style = 2
    chart.width = 18
    chart.height = 9

    data = Reference(ws, min_col=2, max_col=3, min_row=data_start_row, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    s0 = chart.series[0]
    s0.graphicalProperties.solidFill = COLOR_LINE
    s0.graphicalProperties.line.solidFill = COLOR_LINE
    if len(chart.series) > 1:
        s1 = chart.series[1]
        s1.graphicalProperties.solidFill = COLOR_AREA
        s1.graphicalProperties.line.solidFill = COLOR_AREA

    chart.legend.position = 'b'
    ws.add_chart(chart, chart_anchor)


def style_header_row(ws, row, num_cols, fill=HEADER_FILL):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BOTTOM_BORDER


def auto_width(ws, min_width=10, max_width=35):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def add_table(ws, ref, name):
    safe_name = name.replace(" ", "_").replace("-", "_").replace(".", "_")
    table = Table(displayName=safe_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def write_kpi(ws, row, col, label, value, fmt=RUB_FMT):
    """Write a KPI block: small label on top, big value below."""
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = KPI_LABEL_FONT
    value_cell = ws.cell(row=row + 1, column=col, value=value)
    value_cell.font = KPI_FONT
    value_cell.number_format = fmt
    return value_cell


def make_line_chart(title, width=16, height=9):
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.width = width
    chart.height = height
    chart.legend.position = 'b'
    return chart


def make_bar_chart(title, width=16, height=9):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 2
    chart.width = width
    chart.height = height
    chart.legend.position = 'b'
    return chart


def make_pie_chart(title, width=12, height=9):
    chart = PieChart()
    chart.title = title
    chart.style = 2
    chart.width = width
    chart.height = height
    dl = DataLabelList()
    dl.showPercent = True
    dl.showVal = False
    chart.dataLabels = dl
    return chart


# ══════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════════════

def build_dashboard(wb, all_positions, all_operations, portfolio_history,
                    total_xirr, total_deposits, total_withdrawals, divs_by_ticker,
                    total_realized_pnl=0):
    """Лист 1: Дашборд — executive summary."""
    ws = wb.create_sheet("Дашборд")

    # Title
    ws["A1"] = "Портфель Т-Инвестиций"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = datetime.now().strftime('%d.%m.%Y')
    ws["A2"].font = Font(color="808080", size=9)

    # ── KPI Block ──
    total_value = sum(p.get("total_value", 0) for p in all_positions)
    total_pnl = sum(p.get("pnl", 0) for p in all_positions)
    total_divs = sum(divs_by_ticker.values())
    t12m_divs = trailing_12m_income(all_operations)

    row = 4
    write_kpi(ws, row, 1, "СТОИМОСТЬ ПОРТФЕЛЯ", total_value, RUB_FMT_INT)
    write_kpi(ws, row, 3, "P&L (НЕРЕАЛИЗОВАННЫЙ)", total_pnl, RUB_FMT_INT)
    write_kpi(ws, row, 5, "XIRR (ГОДОВАЯ ДОХОДНОСТЬ)", total_xirr if total_xirr else "N/A",
              PCT_FMT if total_xirr else "@")
    write_kpi(ws, row, 7, "ДОХОДЫ ЗА 12 МЕС", t12m_divs, RUB_FMT_INT)

    row = 7
    write_kpi(ws, row, 1, "ВНЕСЕНО", total_deposits, RUB_FMT_INT)
    write_kpi(ws, row, 3, "ВЫВЕДЕНО", total_withdrawals, RUB_FMT_INT)
    write_kpi(ws, row, 5, "РЕАЛИЗОВАННЫЙ P&L", total_realized_pnl, RUB_FMT_INT)
    write_kpi(ws, row, 7, "ДИВИДЕНДЫ + КУПОНЫ", total_divs, RUB_FMT_INT)

    # Color KPIs — subtle
    for r in (5, 8):
        for c in (1, 3, 5, 7):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                if cell.value > 0:
                    cell.font = Font(bold=True, size=13, color=COLOR_POS)
                elif cell.value < 0:
                    cell.font = Font(bold=True, size=13, color=COLOR_NEG)

    # ── Portfolio value chart (compact, row 10) ──
    if portfolio_history:
        add_portfolio_chart(ws, portfolio_history, 50, "A10", "Стоимость портфеля")

    # ── Allocation pie chart (compact, next to portfolio chart) ──
    type_totals = defaultdict(float)
    for p in all_positions:
        type_totals[p.get("instrument_type", "Другое")] += p.get("total_value", 0)

    if type_totals:
        pie_start = 50
        pie_col = 5
        ws.cell(row=pie_start, column=pie_col, value="Тип актива")
        ws.cell(row=pie_start, column=pie_col + 1, value="Стоимость")

        for i, (t, v) in enumerate(sorted(type_totals.items(), key=lambda x: -x[1])):
            ws.cell(row=pie_start + 1 + i, column=pie_col, value=t)
            ws.cell(row=pie_start + 1 + i, column=pie_col + 1, value=v).number_format = RUB_FMT_INT

        pie = make_pie_chart("Аллокация")
        data = Reference(ws, min_col=pie_col + 1, min_row=pie_start,
                         max_row=pie_start + len(type_totals))
        cats = Reference(ws, min_col=pie_col, min_row=pie_start + 1,
                         max_row=pie_start + len(type_totals))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        ws.add_chart(pie, "E10")

    # ── Monthly income bar chart (compact, row 25) ──
    income_by_month = defaultdict(lambda: {"divs": 0, "coupons": 0})
    for op in all_operations:
        if op.get("date") and op.get("type") in INCOME_TYPES:
            m = op["date"].strftime("%Y-%m")
            if op["type"] == "OPERATION_TYPE_DIVIDEND":
                income_by_month[m]["divs"] += op.get("payment", 0)
            else:
                income_by_month[m]["coupons"] += op.get("payment", 0)

    if income_by_month:
        inc_start = 50
        inc_col = 8
        ws.cell(row=inc_start, column=inc_col, value="Месяц")
        ws.cell(row=inc_start, column=inc_col + 1, value="Дивиденды")
        ws.cell(row=inc_start, column=inc_col + 2, value="Купоны")

        for i, (m, vals) in enumerate(sorted(income_by_month.items())):
            ws.cell(row=inc_start + 1 + i, column=inc_col, value=m)
            ws.cell(row=inc_start + 1 + i, column=inc_col + 1, value=vals["divs"]).number_format = RUB_FMT_INT
            ws.cell(row=inc_start + 1 + i, column=inc_col + 2, value=vals["coupons"]).number_format = RUB_FMT_INT

        last_r = inc_start + len(income_by_month)
        bar = make_bar_chart("Доходы по месяцам")
        bar.grouping = "stacked"
        data = Reference(ws, min_col=inc_col + 1, max_col=inc_col + 2,
                         min_row=inc_start, max_row=last_r)
        cats = Reference(ws, min_col=inc_col, min_row=inc_start + 1, max_row=last_r)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        if bar.series:
            bar.series[0].graphicalProperties.solidFill = COLOR_LINE
        if len(bar.series) > 1:
            bar.series[1].graphicalProperties.solidFill = COLOR_AREA
        ws.add_chart(bar, "A25")

    # ── Top positions mini-table ──
    top_start = 25
    top_col = 5
    ws.cell(row=top_start, column=top_col, value="Топ-10 позиций").font = SUBTITLE_FONT

    headers = ["Тикер", "Название", "Стоимость", "P&L %", "Доля %"]
    for j, h in enumerate(headers):
        c = ws.cell(row=top_start + 1, column=top_col + j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    sorted_pos = sorted(all_positions, key=lambda x: abs(x.get("total_value", 0)), reverse=True)[:10]
    for i, p in enumerate(sorted_pos):
        r = top_start + 2 + i
        ws.cell(row=r, column=top_col, value=p.get("ticker", "")).font = Font(bold=True)
        ws.cell(row=r, column=top_col + 1, value=p.get("name", ""))
        ws.cell(row=r, column=top_col + 2, value=p.get("total_value", 0)).number_format = RUB_FMT_INT
        pnl_cell = ws.cell(row=r, column=top_col + 3, value=p.get("pnl_pct", 0))
        pnl_cell.number_format = PCT_FMT
        share = p.get("total_value", 0) / total_value if total_value else 0
        ws.cell(row=r, column=top_col + 4, value=share).number_format = PCT_FMT

    # DataBar on share %
    if sorted_pos:
        share_range = f"{get_column_letter(top_col + 4)}{top_start + 2}:{get_column_letter(top_col + 4)}{top_start + 1 + len(sorted_pos)}"
        ws.conditional_formatting.add(share_range,
            DataBarRule(start_type='min', end_type='max', color='B0BEC5'))

        pnl_range = f"{get_column_letter(top_col + 3)}{top_start + 2}:{get_column_letter(top_col + 3)}{top_start + 1 + len(sorted_pos)}"
        add_pnl_color_scale(ws, pnl_range)

    return ws


def build_returns_sheet(wb, portfolio_history, all_operations, total_xirr, account_xirrs):
    """Лист 2: Доходность — XIRR, heatmap, графики."""
    ws = wb.create_sheet("Доходность")
    ws["A1"] = "Анализ доходности"
    ws["A1"].font = TITLE_FONT

    # ── XIRR by account ──
    ws["A3"] = "XIRR по счетам"
    ws["A3"].font = SUBTITLE_FONT

    headers = ["Счёт", "XIRR"]
    ws.append([])  # row 4
    for j, h in enumerate(headers):
        ws.cell(row=4, column=j + 1, value=h)
    style_header_row(ws, 4, len(headers))

    row = 5
    for acc, rate in sorted(account_xirrs.items()):
        ws.cell(row=row, column=1, value=acc)
        c = ws.cell(row=row, column=2, value=rate if rate else "N/A")
        if rate:
            c.number_format = PCT_FMT
        row += 1

    ws.cell(row=row, column=1, value="ИТОГО").font = Font(bold=True)
    c = ws.cell(row=row, column=2, value=total_xirr if total_xirr else "N/A")
    if total_xirr:
        c.number_format = PCT_FMT
    c.font = Font(bold=True)

    # ── Monthly returns heatmap ──
    if portfolio_history and len(portfolio_history) > 1:
        hm_start = row + 3
        ws.cell(row=hm_start, column=1, value="Месячная доходность (heatmap)").font = SUBTITLE_FONT

        monthly_returns = calculate_monthly_returns(portfolio_history)

        if monthly_returns:
            years = sorted(set(y for y, m in monthly_returns.keys()))
            months_names = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
                            "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]

            # Header row
            hm_header = hm_start + 1
            ws.cell(row=hm_header, column=1, value="Год")
            for m in range(12):
                ws.cell(row=hm_header, column=m + 2, value=months_names[m])
            ws.cell(row=hm_header, column=14, value="Итого")
            style_header_row(ws, hm_header, 14, HEADER_FILL_PURPLE)

            for i, year in enumerate(years):
                r = hm_header + 1 + i
                ws.cell(row=r, column=1, value=year).font = Font(bold=True)
                year_total = 1.0
                for m in range(1, 13):
                    ret = monthly_returns.get((year, m))
                    if ret is not None:
                        ws.cell(row=r, column=m + 1, value=ret).number_format = PCT_FMT
                        year_total *= (1 + ret)
                ws.cell(row=r, column=14, value=year_total - 1).number_format = PCT_FMT

            # Apply ColorScale to heatmap
            hm_range = f"B{hm_header + 1}:N{hm_header + len(years)}"
            add_pnl_color_scale(ws, hm_range)

            # ── Cumulative return chart ──
            cum_start = hm_header + len(years) + 3
            ws.cell(row=cum_start, column=1, value="Дата")
            ws.cell(row=cum_start, column=2, value="Кумулятивная доходность")

            cum_return = 0.0
            cum_data = []
            for i in range(1, len(portfolio_history)):
                d = portfolio_history[i]["date"]
                ret = monthly_returns.get((d.year, d.month), 0)
                cum_return = (1 + cum_return) * (1 + ret) - 1
                cum_data.append((d, cum_return))

            for i, (d, cr) in enumerate(cum_data):
                r = cum_start + 1 + i
                ws.cell(row=r, column=1, value=d).number_format = DATE_FMT
                ws.cell(row=r, column=2, value=cr).number_format = PCT_FMT

            if cum_data:
                chart = make_line_chart("Кумулятивная доходность портфеля")
                data = Reference(ws, min_col=2, min_row=cum_start, max_row=cum_start + len(cum_data))
                cats = Reference(ws, min_col=1, min_row=cum_start + 1, max_row=cum_start + len(cum_data))
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                if chart.series:
                    chart.series[0].graphicalProperties.line.solidFill = COLOR_LINE
                ws.add_chart(chart, f"D{hm_start}")

            # ── Drawdown chart ──
            dd_start = cum_start + len(cum_data) + 3
            ws.cell(row=dd_start, column=1, value="Дата")
            ws.cell(row=dd_start, column=2, value="Просадка от максимума")

            drawdowns = calculate_drawdowns(portfolio_history)
            for i, h in enumerate(portfolio_history):
                r = dd_start + 1 + i
                ws.cell(row=r, column=1, value=h["date"]).number_format = DATE_FMT
                ws.cell(row=r, column=2, value=drawdowns[i]).number_format = PCT_FMT

            if portfolio_history:
                dd_chart = AreaChart()
                dd_chart.title = "Просадка от максимума"
                dd_chart.style = 2
                dd_chart.width = 16
                dd_chart.height = 9
                data = Reference(ws, min_col=2, min_row=dd_start,
                                 max_row=dd_start + len(portfolio_history))
                cats = Reference(ws, min_col=1, min_row=dd_start + 1,
                                 max_row=dd_start + len(portfolio_history))
                dd_chart.add_data(data, titles_from_data=True)
                dd_chart.set_categories(cats)
                if dd_chart.series:
                    dd_chart.series[0].graphicalProperties.solidFill = "CFD8DC"
                    dd_chart.series[0].graphicalProperties.line.solidFill = "78909C"
                ws.add_chart(dd_chart, f"D{cum_start + len(cum_data) + 2}")

    auto_width(ws)
    # no tab color
    return ws


def build_cashflows_sheet(wb, all_operations, portfolio_history):
    """Лист 3: Денежные потоки."""
    ws = wb.create_sheet("Денежные потоки")
    ws["A1"] = "Денежные потоки"
    ws["A1"].font = TITLE_FONT

    monthly = build_monthly_cashflows(all_operations)

    headers = ["Месяц", "Пополнения", "Выводы", "Дивиденды", "Купоны",
               "Комиссии", "Налоги", "Нетто"]
    for j, h in enumerate(headers):
        ws.cell(row=3, column=j + 1, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    for month, vals in monthly.items():
        netto = (vals["deposits"] + vals["withdrawals"] + vals["dividends"] +
                 vals["coupons"] + vals["commissions"] + vals["taxes"])
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=vals["deposits"]).number_format = RUB_FMT
        ws.cell(row=row, column=3, value=vals["withdrawals"]).number_format = RUB_FMT
        ws.cell(row=row, column=4, value=vals["dividends"]).number_format = RUB_FMT
        ws.cell(row=row, column=5, value=vals["coupons"]).number_format = RUB_FMT
        ws.cell(row=row, column=6, value=vals["commissions"]).number_format = RUB_FMT
        ws.cell(row=row, column=7, value=vals["taxes"]).number_format = RUB_FMT
        ws.cell(row=row, column=8, value=netto).number_format = RUB_FMT
        row += 1

    if row > 4:
        add_table(ws, f"A3:{get_column_letter(len(headers))}{row - 1}", "CashFlows")

        # Netto column color scale
        add_pnl_color_scale(ws, f"H4:H{row - 1}")

    # ── Cumulative deposits vs portfolio value chart ──
    if portfolio_history and len(portfolio_history) > 1:
        ch_start = row + 2
        add_portfolio_chart(ws, portfolio_history, ch_start, f"A{ch_start}",
                            "Стоимость портфеля vs Внесённые средства")

    # ── Year-over-year summary ──
    yearly = defaultdict(lambda: {"deposits": 0, "withdrawals": 0, "divs": 0, "coupons": 0, "comms": 0, "taxes": 0})
    for month, vals in monthly.items():
        year = month[:4]
        yearly[year]["deposits"] += vals["deposits"]
        yearly[year]["withdrawals"] += vals["withdrawals"]
        yearly[year]["divs"] += vals["dividends"]
        yearly[year]["coupons"] += vals["coupons"]
        yearly[year]["comms"] += vals["commissions"]
        yearly[year]["taxes"] += vals["taxes"]

    if yearly:
        yoy_start = row + 2 + (len(portfolio_history) + 5 if portfolio_history else 0)
        ws.cell(row=yoy_start, column=1, value="Сводка по годам").font = SUBTITLE_FONT
        yoy_headers = ["Год", "Пополнения", "Выводы", "Дивиденды", "Купоны", "Комиссии", "Налоги"]
        for j, h in enumerate(yoy_headers):
            ws.cell(row=yoy_start + 1, column=j + 1, value=h)
        style_header_row(ws, yoy_start + 1, len(yoy_headers))

        r = yoy_start + 2
        for year, vals in sorted(yearly.items()):
            ws.cell(row=r, column=1, value=year)
            ws.cell(row=r, column=2, value=vals["deposits"]).number_format = RUB_FMT
            ws.cell(row=r, column=3, value=vals["withdrawals"]).number_format = RUB_FMT
            ws.cell(row=r, column=4, value=vals["divs"]).number_format = RUB_FMT
            ws.cell(row=r, column=5, value=vals["coupons"]).number_format = RUB_FMT
            ws.cell(row=r, column=6, value=vals["comms"]).number_format = RUB_FMT
            ws.cell(row=r, column=7, value=vals["taxes"]).number_format = RUB_FMT
            r += 1

    auto_width(ws)
    ws.freeze_panes = "A4"
    # no tab color
    return ws


def build_portfolio_sheet(wb, all_positions, divs_by_ticker):
    """Лист 4: Портфель — текущие позиции с расширенной аналитикой."""
    ws = wb.create_sheet("Портфель")

    headers = [
        "Счёт", "Тикер", "Название", "Тип", "Сектор", "Валюта",
        "Кол-во", "Средняя цена", "Средняя FIFO", "Текущая цена",
        "Стоимость", "P&L ₽", "P&L %", "Доля %",
        "Дивиденды получено", "Див. доходность",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    total_value = sum(p.get("total_value", 0) for p in all_positions)

    for pos in sorted(all_positions, key=lambda x: abs(x.get("total_value", 0)), reverse=True):
        ticker = pos.get("ticker", "")
        divs = divs_by_ticker.get(ticker, 0)
        cost_basis = pos.get("quantity", 0) * pos.get("avg_price", 0)
        div_yield = divs / cost_basis if cost_basis > 0 else 0
        share = pos.get("total_value", 0) / total_value if total_value else 0

        ws.append([
            pos.get("account_name", ""),
            ticker,
            pos.get("name", ""),
            pos.get("instrument_type", ""),
            pos.get("sector", ""),
            pos.get("currency", ""),
            pos.get("quantity", 0),
            pos.get("avg_price", 0),
            pos.get("avg_price_fifo", pos.get("avg_price", 0)),
            pos.get("cur_price", 0),
            pos.get("total_value", 0),
            pos.get("pnl", 0),
            pos.get("pnl_pct", 0),
            share,
            divs,
            div_yield,
        ])

    max_row = ws.max_row
    # Formatting
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            cell.border = THIN_BORDER
        row[7].number_format = RUB_FMT   # avg price
        row[8].number_format = RUB_FMT   # avg fifo
        row[9].number_format = RUB_FMT   # cur price
        row[10].number_format = RUB_FMT  # value
        row[11].number_format = RUB_FMT  # pnl
        row[12].number_format = PCT_FMT  # pnl %
        row[13].number_format = PCT_FMT  # share %
        row[14].number_format = RUB_FMT  # divs
        row[15].number_format = PCT_FMT  # div yield

    if max_row > 1:
        add_table(ws, f"A1:{get_column_letter(len(headers))}{max_row}", "Portfolio")

        # ColorScale on P&L %
        add_pnl_color_scale(ws, f"M2:M{max_row}")

        # DataBar on share %
        ws.conditional_formatting.add(f"N2:N{max_row}",
            DataBarRule(start_type='min', end_type='max', color='B0BEC5'))

        # DataBar on value
        ws.conditional_formatting.add(f"K2:K{max_row}",
            DataBarRule(start_type='min', end_type='max', color='B0BEC5'))

    auto_width(ws)
    ws.freeze_panes = "C2"
    # no tab color — clean
    return ws


def build_closed_positions_sheet(wb, closed_lots):
    """Лист 5: Закрытые позиции — реализованный P&L."""
    ws = wb.create_sheet("Закрытые позиции")
    ws["A1"] = "Реализованный P&L (FIFO)"
    ws["A1"].font = TITLE_FONT

    headers = [
        "Тикер", "Название", "Счёт", "Дата покупки", "Дата продажи",
        "Кол-во", "Цена покупки", "Цена продажи",
        "P&L ₽", "P&L %", "Дней в позиции", "Годовая доходность",
    ]
    for j, h in enumerate(headers):
        ws.cell(row=3, column=j + 1, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    total_realized = 0
    for lot in sorted(closed_lots, key=lambda x: x.get("sell_date") or datetime.min, reverse=True):
        ws.cell(row=row, column=1, value=lot["ticker"])
        ws.cell(row=row, column=2, value=lot["name"])
        ws.cell(row=row, column=3, value=lot["account"])
        ws.cell(row=row, column=4, value=lot["buy_date"]).number_format = DATE_FMT
        ws.cell(row=row, column=5, value=lot["sell_date"]).number_format = DATE_FMT
        ws.cell(row=row, column=6, value=lot["quantity"])
        ws.cell(row=row, column=7, value=lot["buy_price"]).number_format = RUB_FMT
        ws.cell(row=row, column=8, value=lot["sell_price"]).number_format = RUB_FMT
        ws.cell(row=row, column=9, value=lot["pnl"]).number_format = RUB_FMT
        ws.cell(row=row, column=10, value=lot["pnl_pct"]).number_format = PCT_FMT
        ws.cell(row=row, column=11, value=lot["days_held"])
        ws.cell(row=row, column=12, value=lot["annual_return"]).number_format = PCT_FMT
        total_realized += lot["pnl"]
        row += 1

    # Total row
    if closed_lots:
        ws.cell(row=row, column=1, value="ИТОГО").font = Font(bold=True)
        tc = ws.cell(row=row, column=9, value=total_realized)
        tc.number_format = RUB_FMT
        tc.font = Font(bold=True)

        add_table(ws, f"A3:{get_column_letter(len(headers))}{row - 1}", "ClosedPositions")

        # ColorScale on P&L
        add_pnl_color_scale(ws, f"I4:I{row - 1}")

        # Bar chart: top winners/losers
        top_lots = sorted(closed_lots, key=lambda x: x["pnl"], reverse=True)
        chart_lots = top_lots[:5] + sorted(closed_lots, key=lambda x: x["pnl"])[:5]

        ch_row = row + 3
        ws.cell(row=ch_row, column=1, value="Тикер")
        ws.cell(row=ch_row, column=2, value="P&L")
        for i, lot in enumerate(chart_lots):
            ws.cell(row=ch_row + 1 + i, column=1, value=f"{lot['ticker']} ({lot['sell_date'].strftime('%d.%m') if lot['sell_date'] else ''})")
            ws.cell(row=ch_row + 1 + i, column=2, value=lot["pnl"]).number_format = RUB_FMT

        bar = make_bar_chart("Топ P&L")
        data = Reference(ws, min_col=2, min_row=ch_row, max_row=ch_row + len(chart_lots))
        cats = Reference(ws, min_col=1, min_row=ch_row + 1, max_row=ch_row + len(chart_lots))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "E3")

        # ── Summary by ticker ──
        sum_start = row + 3 + len(chart_lots) + 3
        ws.cell(row=sum_start, column=1, value="Сводка по тикерам").font = SUBTITLE_FONT

        sum_headers = ["Тикер", "Сделок", "Объём покупок", "Объём продаж", "Реализованный P&L", "Ср. дней"]
        for j, h in enumerate(sum_headers):
            ws.cell(row=sum_start + 1, column=j + 1, value=h)
        style_header_row(ws, sum_start + 1, len(sum_headers))

        ticker_summary = defaultdict(lambda: {"count": 0, "buy_vol": 0, "sell_vol": 0, "pnl": 0, "days": []})
        for lot in closed_lots:
            t = lot["ticker"]
            ticker_summary[t]["count"] += 1
            ticker_summary[t]["buy_vol"] += lot["quantity"] * lot["buy_price"]
            ticker_summary[t]["sell_vol"] += lot["quantity"] * lot["sell_price"]
            ticker_summary[t]["pnl"] += lot["pnl"]
            ticker_summary[t]["days"].append(lot["days_held"])

        sr = sum_start + 2
        for t in sorted(ticker_summary.keys(), key=lambda x: abs(ticker_summary[x]["pnl"]), reverse=True):
            s = ticker_summary[t]
            ws.cell(row=sr, column=1, value=t).font = Font(bold=True)
            ws.cell(row=sr, column=2, value=s["count"])
            ws.cell(row=sr, column=3, value=s["buy_vol"]).number_format = RUB_FMT
            ws.cell(row=sr, column=4, value=s["sell_vol"]).number_format = RUB_FMT
            pnl_cell = ws.cell(row=sr, column=5, value=s["pnl"])
            pnl_cell.number_format = RUB_FMT
            pnl_cell.font = Font(color=COLOR_POS if s["pnl"] >= 0 else COLOR_NEG)
            avg_days = sum(s["days"]) / len(s["days"]) if s["days"] else 0
            ws.cell(row=sr, column=6, value=round(avg_days))
            sr += 1

    auto_width(ws)
    # no tab color
    return ws


def build_dividends_sheet(wb, all_operations):
    """Лист 6: Дивиденды и купоны — с heatmap и аналитикой."""
    ws = wb.create_sheet("Дивиденды и купоны")
    ws["A1"] = "Дивиденды и купоны"
    ws["A1"].font = TITLE_FONT

    # Trailing 12m income
    div_ops = [op for op in all_operations if op.get("type") in INCOME_TYPES]
    t12m_total = trailing_12m_income(all_operations)

    ws["A2"] = f"Доход за последние 12 месяцев: "
    ws["A2"].font = Font(color="808080")
    ws.cell(row=2, column=4, value=t12m_total).number_format = RUB_FMT
    ws.cell(row=2, column=4).font = Font(bold=True, size=11, color=COLOR_POS)

    headers = ["Дата", "Счёт", "Тикер", "Название", "Тип", "Сумма", "Валюта", "Год", "Квартал"]
    for j, h in enumerate(headers):
        ws.cell(row=4, column=j + 1, value=h)
    style_header_row(ws, 4, len(headers), HEADER_FILL_GREEN)

    row = 5
    for op in sorted(div_ops, key=lambda x: x.get("date") or datetime.min, reverse=True):
        d = op.get("date")
        ws.cell(row=row, column=1, value=d).number_format = DATETIME_FMT
        ws.cell(row=row, column=2, value=op.get("account_name", ""))
        ws.cell(row=row, column=3, value=op.get("ticker", ""))
        ws.cell(row=row, column=4, value=op.get("name", ""))
        ws.cell(row=row, column=5, value=op.get("type_display", ""))
        ws.cell(row=row, column=6, value=op.get("payment", 0)).number_format = RUB_FMT
        ws.cell(row=row, column=7, value=op.get("currency", ""))
        ws.cell(row=row, column=8, value=d.year if d else "")
        ws.cell(row=row, column=9, value=f"Q{(d.month - 1) // 3 + 1}" if d else "")
        row += 1

    if row > 5:
        add_table(ws, f"A4:{get_column_letter(len(headers))}{row - 1}", "DividendsCoupons")

    # ── Heatmap: ticker × month ──
    hm_start = row + 2
    ws.cell(row=hm_start, column=1, value="Календарь доходов (тикер × месяц)").font = SUBTITLE_FONT

    income_heatmap = defaultdict(lambda: defaultdict(float))
    for op in div_ops:
        d = op.get("date")
        if d:
            income_heatmap[op.get("ticker", "")][d.strftime("%Y-%m")] += op.get("payment", 0)

    if income_heatmap:
        all_months = sorted(set(m for ticker_months in income_heatmap.values() for m in ticker_months))
        tickers = sorted(income_heatmap.keys())

        # Header
        ws.cell(row=hm_start + 1, column=1, value="Тикер")
        for j, m in enumerate(all_months):
            ws.cell(row=hm_start + 1, column=j + 2, value=m)
        ws.cell(row=hm_start + 1, column=len(all_months) + 2, value="Итого")
        style_header_row(ws, hm_start + 1, len(all_months) + 2, HEADER_FILL_GREEN)

        for i, ticker in enumerate(tickers):
            r = hm_start + 2 + i
            ws.cell(row=r, column=1, value=ticker).font = Font(bold=True)
            total = 0
            for j, m in enumerate(all_months):
                val = income_heatmap[ticker].get(m, 0)
                if val > 0:
                    ws.cell(row=r, column=j + 2, value=val).number_format = RUB_FMT_INT
                total += val
            ws.cell(row=r, column=len(all_months) + 2, value=total).number_format = RUB_FMT

        # ColorScale
        hm_range = f"B{hm_start + 2}:{get_column_letter(len(all_months) + 1)}{hm_start + 1 + len(tickers)}"
        ws.conditional_formatting.add(hm_range,
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FFFFFF',
                end_type='max', end_color='63BE7B'))

    # ── Annual income bar chart ──
    annual_income = defaultdict(float)
    for op in div_ops:
        d = op.get("date")
        if d:
            annual_income[d.year] += op.get("payment", 0)

    if annual_income:
        ch_start = hm_start + len(income_heatmap) + 5
        ws.cell(row=ch_start, column=1, value="Год")
        ws.cell(row=ch_start, column=2, value="Доход")
        for i, (year, total) in enumerate(sorted(annual_income.items())):
            ws.cell(row=ch_start + 1 + i, column=1, value=str(year))
            ws.cell(row=ch_start + 1 + i, column=2, value=total).number_format = RUB_FMT

        bar = make_bar_chart("Доходы по годам", 20, 12)
        data = Reference(ws, min_col=2, min_row=ch_start, max_row=ch_start + len(annual_income))
        cats = Reference(ws, min_col=1, min_row=ch_start + 1, max_row=ch_start + len(annual_income))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        if bar.series:
            bar.series[0].graphicalProperties.solidFill = COLOR_LINE
        ws.add_chart(bar, f"D{hm_start + len(income_heatmap) + 4}")

    auto_width(ws, max_width=14)
    ws.freeze_panes = "A5"
    # no tab color
    return ws


def build_taxes_sheet(wb, all_operations, all_positions):
    """Лист 7: Налоги."""
    ws = wb.create_sheet("Налоги")
    ws["A1"] = "Налоговая аналитика"
    ws["A1"].font = TITLE_FONT

    tax_data = aggregate_taxes(all_operations)

    # ── Taxes by year ──
    ws["A3"] = "Налоги уплаченные по годам"
    ws["A3"].font = SUBTITLE_FONT

    all_tax_types = sorted(set(t for year_taxes in tax_data.values() for t in year_taxes))
    headers = ["Год"] + all_tax_types + ["Итого"]
    for j, h in enumerate(headers):
        ws.cell(row=4, column=j + 1, value=h)
    style_header_row(ws, 4, len(headers), HEADER_FILL_RED)

    row = 5
    for year in sorted(tax_data.keys()):
        ws.cell(row=row, column=1, value=year)
        total = 0
        for j, tt in enumerate(all_tax_types):
            val = tax_data[year].get(tt, 0)
            ws.cell(row=row, column=j + 2, value=val).number_format = RUB_FMT
            total += val
        ws.cell(row=row, column=len(all_tax_types) + 2, value=total).number_format = RUB_FMT
        ws.cell(row=row, column=len(all_tax_types) + 2).font = Font(bold=True)
        row += 1

    # ── Potential tax on unrealized gains ──
    unreal_start = row + 2
    ws.cell(row=unreal_start, column=1, value="Потенциальный налог на нереализованный P&L").font = SUBTITLE_FONT

    total_unrealized = sum(p.get("pnl", 0) for p in all_positions if p.get("pnl", 0) > 0)
    potential_tax = total_unrealized * 0.13

    ws.cell(row=unreal_start + 1, column=1, value="Нереализованная прибыль (только положительная)")
    ws.cell(row=unreal_start + 1, column=2, value=total_unrealized).number_format = RUB_FMT

    ws.cell(row=unreal_start + 2, column=1, value="Потенциальный НДФЛ (13%)")
    ws.cell(row=unreal_start + 2, column=2, value=potential_tax).number_format = RUB_FMT
    ws.cell(row=unreal_start + 2, column=2).font = Font(bold=True, color=COLOR_NEG)

    # ── IIS tracker ──
    iis_start = unreal_start + 5
    ws.cell(row=iis_start, column=1, value="Трекер ИИС (тип А — вычет на взносы)").font = SUBTITLE_FONT

    iis_deposits = defaultdict(float)
    for op in all_operations:
        if op.get("type") in DEPOSIT_TYPES:
            d = op.get("date")
            acc = op.get("account_name", "")
            if d and "ИИС" in acc:
                iis_deposits[(d.year, acc)] += op.get("payment", 0)

    if iis_deposits:
        ws.cell(row=iis_start + 1, column=1, value="Год")
        ws.cell(row=iis_start + 1, column=2, value="Счёт")
        ws.cell(row=iis_start + 1, column=3, value="Внесено")
        ws.cell(row=iis_start + 1, column=4, value="База для вычета (макс 400к)")
        ws.cell(row=iis_start + 1, column=5, value="Потенциальный вычет (13%)")
        style_header_row(ws, iis_start + 1, 5, HEADER_FILL_GREEN)

        r = iis_start + 2
        for (year, acc), amount in sorted(iis_deposits.items()):
            ws.cell(row=r, column=1, value=year)
            ws.cell(row=r, column=2, value=acc)
            ws.cell(row=r, column=3, value=amount).number_format = RUB_FMT
            base = min(amount, 400_000)
            ws.cell(row=r, column=4, value=base).number_format = RUB_FMT
            deduction = base * 0.13
            ws.cell(row=r, column=5, value=deduction).number_format = RUB_FMT
            ws.cell(row=r, column=5).font = Font(bold=True, color=COLOR_POS)
            r += 1

    auto_width(ws)
    # no tab color
    return ws


def build_accounts_sheet(wb, all_positions, all_operations, account_xirrs):
    """Лист 8: По счетам — сводка."""
    ws = wb.create_sheet("По счетам")
    ws["A1"] = "Аналитика по счетам"
    ws["A1"].font = TITLE_FONT

    headers = [
        "Счёт", "Тип", "Стоимость", "P&L",
        "XIRR", "Внесено", "Выведено",
        "Дивиденды", "Купоны", "Комиссии", "Налоги",
    ]
    for j, h in enumerate(headers):
        ws.cell(row=3, column=j + 1, value=h)
    style_header_row(ws, 3, len(headers))

    account_data = {}
    for p in all_positions:
        acc = p.get("account_name", "")
        if acc not in account_data:
            account_data[acc] = {
                "type": p.get("account_type_display", ""),
                "value": 0, "pnl": 0, "deposits": 0, "withdrawals": 0,
                "divs": 0, "coupons": 0, "commissions": 0, "taxes": 0,
            }
        account_data[acc]["value"] += p.get("total_value", 0)
        account_data[acc]["pnl"] += p.get("pnl", 0)

    for op in all_operations:
        acc = op.get("account_name", "")
        if acc not in account_data:
            account_data[acc] = {
                "type": "", "value": 0, "pnl": 0, "deposits": 0, "withdrawals": 0,
                "divs": 0, "coupons": 0, "commissions": 0, "taxes": 0,
            }
        op_type = op.get("type", "")
        payment = op.get("payment", 0)

        if op_type in DEPOSIT_TYPES:
            account_data[acc]["deposits"] += payment
        elif op_type in WITHDRAWAL_TYPES:
            account_data[acc]["withdrawals"] += payment
        elif op_type == "OPERATION_TYPE_DIVIDEND":
            account_data[acc]["divs"] += payment
        elif op_type == "OPERATION_TYPE_COUPON":
            account_data[acc]["coupons"] += payment
        elif op_type in COMMISSION_TYPES:
            account_data[acc]["commissions"] += payment
        elif op_type in TAX_TYPES:
            account_data[acc]["taxes"] += payment

    row = 4
    for acc in sorted(account_data.keys()):
        d = account_data[acc]
        xirr_val = account_xirrs.get(acc)
        ws.cell(row=row, column=1, value=acc).font = Font(bold=True)
        ws.cell(row=row, column=2, value=d["type"])
        ws.cell(row=row, column=3, value=d["value"]).number_format = RUB_FMT
        ws.cell(row=row, column=4, value=d["pnl"]).number_format = RUB_FMT
        c = ws.cell(row=row, column=5, value=xirr_val if xirr_val else "N/A")
        if xirr_val:
            c.number_format = PCT_FMT
        ws.cell(row=row, column=6, value=d["deposits"]).number_format = RUB_FMT
        ws.cell(row=row, column=7, value=d["withdrawals"]).number_format = RUB_FMT
        ws.cell(row=row, column=8, value=d["divs"]).number_format = RUB_FMT
        ws.cell(row=row, column=9, value=d["coupons"]).number_format = RUB_FMT
        ws.cell(row=row, column=10, value=d["commissions"]).number_format = RUB_FMT
        ws.cell(row=row, column=11, value=d["taxes"]).number_format = RUB_FMT
        row += 1

    if row > 4:
        # ColorScale on P&L
        add_pnl_color_scale(ws, f"D4:D{row - 1}")

        # Bar chart comparing accounts
        bar = make_bar_chart("Стоимость портфеля по счетам", 24, 14)
        data = Reference(ws, min_col=3, min_row=3, max_row=row - 1)
        cats = Reference(ws, min_col=1, min_row=4, max_row=row - 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        if bar.series:
            bar.series[0].graphicalProperties.solidFill = COLOR_LINE
        ws.add_chart(bar, "A" + str(row + 2))

    auto_width(ws)
    return ws


def build_operations_sheet(wb, all_operations):
    """Лист 9: Операции — с группой и running balance."""
    ws = wb.create_sheet("Операции")

    headers = [
        "Дата", "Счёт", "Тикер", "Название", "Тип операции", "Группа",
        "Кол-во", "Цена", "Сумма", "Валюта", "Комиссия", "Баланс (куммул.)",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    running_balance = 0
    for op in all_operations:  # already sorted in main()
        payment = op.get("payment", 0)
        running_balance += payment
        group = OPERATION_GROUPS.get(op.get("type", ""), "Прочее")

        ws.append([
            op.get("date"),
            op.get("account_name", ""),
            op.get("ticker", ""),
            op.get("name", ""),
            op.get("type_display", ""),
            group,
            op.get("quantity", ""),
            op.get("price", ""),
            payment,
            op.get("currency", ""),
            op.get("commission", 0),
            running_balance,
        ])

    max_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        if row[0].value:
            row[0].number_format = DATETIME_FMT
        row[7].number_format = RUB_FMT if row[7].value else "@"
        row[8].number_format = RUB_FMT
        row[10].number_format = RUB_FMT
        row[11].number_format = RUB_FMT

    if max_row > 1:
        add_table(ws, f"A1:{get_column_letter(len(headers))}{max_row}", "Operations")

        # ColorScale on payment
        add_pnl_color_scale(ws, f"I2:I{max_row}")

    auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


def build_commissions_sheet(wb, all_operations):
    """Лист 10: Комиссии."""
    ws = wb.create_sheet("Комиссии")
    ws["A1"] = "Анализ комиссий"
    ws["A1"].font = TITLE_FONT

    headers = ["Месяц", "Счёт", "Комиссия", "Объём торгов", "% от объёма"]
    for j, h in enumerate(headers):
        ws.cell(row=3, column=j + 1, value=h)
    style_header_row(ws, 3, len(headers), HEADER_FILL_ORANGE)

    # Aggregate
    comm_data = defaultdict(lambda: {"commission": 0, "volume": 0})
    for op in all_operations:
        d = op.get("date")
        if not d:
            continue
        key = (d.strftime("%Y-%m"), op.get("account_name", ""))
        comm_data[key]["commission"] += abs(op.get("commission", 0))
        if op.get("type") in BUY_SELL_TYPES:
            comm_data[key]["volume"] += abs(op.get("payment", 0))

    row = 4
    for (month, acc), vals in sorted(comm_data.items()):
        if vals["commission"] == 0:
            continue
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=acc)
        ws.cell(row=row, column=3, value=vals["commission"]).number_format = RUB_FMT
        ws.cell(row=row, column=4, value=vals["volume"]).number_format = RUB_FMT
        pct = vals["commission"] / vals["volume"] if vals["volume"] else 0
        ws.cell(row=row, column=5, value=pct).number_format = '0.000%'
        row += 1

    if row > 4:
        add_table(ws, f"A3:{get_column_letter(len(headers))}{row - 1}", "CommissionsDetail")

        # Bar chart
        # Aggregate by month only
        monthly_comm = defaultdict(float)
        for (month, acc), vals in comm_data.items():
            monthly_comm[month] += vals["commission"]

        ch_start = row + 2
        ws.cell(row=ch_start, column=1, value="Месяц")
        ws.cell(row=ch_start, column=2, value="Комиссия")
        for i, (m, c) in enumerate(sorted(monthly_comm.items())):
            ws.cell(row=ch_start + 1 + i, column=1, value=m)
            ws.cell(row=ch_start + 1 + i, column=2, value=c).number_format = RUB_FMT

        if monthly_comm:
            bar = make_bar_chart("Комиссии по месяцам", 24, 12)
            data = Reference(ws, min_col=2, min_row=ch_start, max_row=ch_start + len(monthly_comm))
            cats = Reference(ws, min_col=1, min_row=ch_start + 1, max_row=ch_start + len(monthly_comm))
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            if bar.series:
                bar.series[0].graphicalProperties.solidFill = COLOR_LINE
            ws.add_chart(bar, "D3")

    auto_width(ws)
    # no tab color
    return ws


def build_instruments_sheet(wb, instruments_cache):
    """Лист 11: Справочник инструментов."""
    ws = wb.create_sheet("Справочник")

    headers = ["UID", "FIGI", "Тикер", "Название", "Тип", "Сектор", "Валюта", "Страна", "Лот"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for uid, inst in sorted(instruments_cache.items(), key=lambda x: x[1].get("ticker", "")):
        ws.append([
            uid,
            inst.get("figi", ""),
            inst.get("ticker", ""),
            inst.get("name", ""),
            INSTRUMENT_TYPE_MAP.get(inst.get("instrumentType", inst.get("instrumentKind", "")), ""),
            inst.get("sector", ""),
            inst.get("currency", "").upper(),
            inst.get("countryOfRisk", ""),
            inst.get("lot", ""),
        ])

    if ws.max_row > 1:
        add_table(ws, f"A1:{get_column_letter(len(headers))}{ws.max_row}", "Instruments")

    auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


def build_position_sheet(wb, ticker, position_info, ticker_operations, candles_data):
    """Лист для отдельной позиции: график цены + маркеры сделок."""
    sheet_name = f"📈 {ticker}"[:31]
    ws = wb.create_sheet(sheet_name)

    # ── Mini-summary ──
    ws["A1"] = ticker
    ws["A1"].font = Font(bold=True, size=12)
    ws["C1"] = position_info.get("name", "")
    ws["C1"].font = Font(size=10, color="808080")

    row = 3
    labels = [
        ("Сектор", position_info.get("sector", "")),
        ("Тип", position_info.get("instrument_type", "")),
        ("Текущая цена", position_info.get("cur_price", 0)),
        ("Средняя цена покупки", position_info.get("avg_price", 0)),
        ("Кол-во", position_info.get("quantity", 0)),
        ("Стоимость", position_info.get("total_value", 0)),
        ("P&L ₽", position_info.get("pnl", 0)),
        ("P&L %", position_info.get("pnl_pct", 0)),
    ]
    for label, val in labels:
        ws.cell(row=row, column=1, value=label).font = Font(color="808080")
        c = ws.cell(row=row, column=2, value=val)
        if "цена" in label.lower() or label in ("Стоимость", "P&L ₽"):
            c.number_format = RUB_FMT
        elif label == "P&L %":
            c.number_format = PCT_FMT
        if label == "P&L ₽":
            c.font = Font(bold=True, color=COLOR_POS if val >= 0 else "C0392B")
        row += 1

    # ── Price data + buy/sell markers ──
    data_start = 13
    ws.cell(row=data_start, column=1, value="Дата")
    ws.cell(row=data_start, column=2, value="Цена закрытия")
    ws.cell(row=data_start, column=3, value="Покупка")
    ws.cell(row=data_start, column=4, value="Продажа")
    style_header_row(ws, data_start, 4)

    # Build buy/sell lookup: date -> (price, qty, type)
    trades_by_date = defaultdict(list)
    for op in ticker_operations:
        d = op.get("date")
        if d and op["type"] in BUY_SELL_TYPES:
            day = d.replace(hour=0, minute=0, second=0, microsecond=0)
            trades_by_date[day].append(op)

    r = data_start + 1
    if candles_data and candles_data.get("dates"):
        for d, close in zip(candles_data["dates"], candles_data["prices"]):
            ws.cell(row=r, column=1, value=d).number_format = DATE_FMT
            ws.cell(row=r, column=2, value=close).number_format = RUB_FMT

            day_key = d.replace(hour=0, minute=0, second=0, microsecond=0)
            for trade in trades_by_date.get(day_key, []):
                price = abs(trade.get("price", 0))
                if price == 0:
                    price = close
                if trade["type"] == "OPERATION_TYPE_BUY":
                    ws.cell(row=r, column=3, value=price).number_format = RUB_FMT
                else:
                    ws.cell(row=r, column=4, value=price).number_format = RUB_FMT

            r += 1

    last_data_row = r - 1

    # ── Chart ──
    if last_data_row > data_start:
        chart = LineChart()
        chart.title = f"{ticker} — цена и сделки"
        chart.style = 2
        chart.width = 20
        chart.height = 10
        chart.legend.position = 'b'

        # Series 1: Price line
        price_data = Reference(ws, min_col=2, min_row=data_start, max_row=last_data_row)
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=last_data_row)
        chart.add_data(price_data, titles_from_data=True)
        chart.set_categories(cats)

        price_series = chart.series[0]
        price_series.graphicalProperties.line.solidFill = COLOR_LINE
        price_series.graphicalProperties.line.width = 22000  # ~2pt
        price_series.marker = None
        price_series.smooth = False

        # Series 2: Buy markers
        buy_data = Reference(ws, min_col=3, min_row=data_start, max_row=last_data_row)
        chart.add_data(buy_data, titles_from_data=True)
        buy_series = chart.series[1]
        buy_series.graphicalProperties.line = LineProperties(w=0)
        buy_series.marker = Marker(symbol='triangle', size=7)
        buy_series.marker.graphicalProperties.solidFill = COLOR_POS
        buy_series.marker.graphicalProperties.line.solidFill = COLOR_POS
        buy_series.smooth = False

        # Series 3: Sell markers
        sell_data = Reference(ws, min_col=4, min_row=data_start, max_row=last_data_row)
        chart.add_data(sell_data, titles_from_data=True)
        sell_series = chart.series[2]
        sell_series.graphicalProperties.line = LineProperties(w=0)
        sell_series.marker = Marker(symbol='diamond', size=7)
        sell_series.marker.graphicalProperties.solidFill = COLOR_NEG
        sell_series.marker.graphicalProperties.line.solidFill = COLOR_NEG
        sell_series.smooth = False

        ws.add_chart(chart, "E3")

    # ── Trades table ──
    trades_start = data_start
    trades_col = 6
    ws.cell(row=trades_start, column=trades_col, value="Дата")
    ws.cell(row=trades_start, column=trades_col + 1, value="Операция")
    ws.cell(row=trades_start, column=trades_col + 2, value="Кол-во")
    ws.cell(row=trades_start, column=trades_col + 3, value="Цена")
    ws.cell(row=trades_start, column=trades_col + 4, value="Сумма")
    style_header_row(ws, trades_start, 5, fill=HEADER_FILL)
    # Adjust: style only the trades header cols


    trade_ops = [op for op in ticker_operations
                 if op["type"] in BUY_SELL_TYPES]
    trade_ops.sort(key=lambda x: x.get("date") or datetime.min, reverse=True)

    for i, op in enumerate(trade_ops):
        tr = trades_start + 1 + i
        ws.cell(row=tr, column=trades_col, value=op.get("date")).number_format = DATE_FMT
        op_name = "Покупка" if op["type"] == "OPERATION_TYPE_BUY" else "Продажа"
        c = ws.cell(row=tr, column=trades_col + 1, value=op_name)
        c.font = Font(color=COLOR_POS if op["type"] == "OPERATION_TYPE_BUY" else "C0392B")
        ws.cell(row=tr, column=trades_col + 2, value=abs(op.get("quantity", 0)))
        ws.cell(row=tr, column=trades_col + 3, value=abs(op.get("price", 0))).number_format = RUB_FMT
        ws.cell(row=tr, column=trades_col + 4, value=op.get("payment", 0)).number_format = RUB_FMT

    auto_width(ws)
    return ws


def build_history_sheet(wb, portfolio_history):
    """Лист: Портфель (история) — месячные снимки."""
    ws = wb.create_sheet("Портфель (история)")

    headers = ["Дата", "Стоимость", "Внесено (куммул.)", "Выведено (куммул.)",
               "Доходность %", "Просадка от макс. %"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    drawdowns = calculate_drawdowns(portfolio_history)
    for i, h in enumerate(portfolio_history):
        invested = h["cum_deposits"] - h["cum_withdrawals"]
        ret = (h["value"] - invested) / invested if invested > 0 else 0

        ws.append([
            h["date"],
            h["value"],
            h["cum_deposits"],
            h["cum_withdrawals"],
            ret,
            drawdowns[i],
        ])

    max_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        row[0].number_format = DATE_FMT
        row[1].number_format = RUB_FMT_INT
        row[2].number_format = RUB_FMT_INT
        row[3].number_format = RUB_FMT_INT
        row[4].number_format = PCT_FMT
        row[5].number_format = PCT_FMT

    if max_row > 1:
        add_table(ws, f"A1:{get_column_letter(len(headers))}{max_row}", "PortfolioHistory")

    auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════

def main():
    print("═" * 60)
    print("  T-Invest Analytics Report Builder v2")
    print("═" * 60)
    print()

    # ── Step 1: Get accounts ──
    print("Подключаюсь к T-Invest API...")
    accounts = get_accounts()
    if not accounts:
        print("Не удалось получить список счетов!")
        sys.exit(1)
    print(f"Найдено {len(accounts)} открытых счетов")

    instruments_cache = {}
    all_positions = []
    all_operations = []
    account_cashflows = defaultdict(list)  # for per-account XIRR

    # ── Step 2: Fetch data per account ──
    for acc in accounts:
        acc_id = acc["id"]
        acc_name = acc.get("name", acc_id)
        acc_type = ACCOUNT_TYPE_MAP.get(acc.get("type", ""), acc.get("type", ""))
        print(f"\n── {acc_name} ({acc_type}) ──")

        # Portfolio
        print("  Загружаю портфель...")
        portfolio = get_portfolio(acc_id)
        if portfolio:
            positions = portfolio.get("positions", [])
            print(f"  Позиций: {len(positions)}")

            for pos in positions:
                uid = pos.get("instrumentUid", "")
                ticker = pos.get("ticker", pos.get("figi", ""))

                if uid and uid not in instruments_cache:
                    inst = get_instrument_by_uid(uid)
                    if inst:
                        instruments_cache[uid] = inst

                inst_info = instruments_cache.get(uid, {})
                name = inst_info.get("name", ticker)
                sector = inst_info.get("sector", "")
                inst_type = INSTRUMENT_TYPE_MAP.get(
                    pos.get("instrumentType", ""), pos.get("instrumentType", "")
                )
                currency = inst_info.get("currency", "rub").upper()

                qty = quotation_value(pos.get("quantity"))
                avg_price = money_value(pos.get("averagePositionPrice"))
                avg_price_fifo = money_value(pos.get("averagePositionPriceFifo")) or avg_price
                cur_price = money_value(pos.get("currentPrice"))
                expected_yield = quotation_value(pos.get("expectedYield"))

                cost_basis = qty * avg_price
                pnl = expected_yield
                pnl_pct = pnl / cost_basis if cost_basis else 0

                all_positions.append({
                    "account_name": acc_name,
                    "account_type_display": acc_type,
                    "ticker": ticker,
                    "figi": pos.get("figi", inst_info.get("figi", "")),
                    "name": name,
                    "instrument_type": inst_type,
                    "sector": sector,
                    "currency": currency,
                    "quantity": qty,
                    "avg_price": avg_price,
                    "avg_price_fifo": avg_price_fifo,
                    "cur_price": cur_price,
                    "total_value": qty * cur_price,
                    "pnl": pnl,
                    "pnl_pct": pnl_pct,
                })

        # Operations
        print("  Загружаю операции...")
        to_date = datetime.now(timezone.utc)
        ops = get_operations(acc_id, ANALYSIS_START, to_date)
        print(f"  Операций: {len(ops)}")

        for op in ops:
            op_type = op.get("type", op.get("operationType", ""))
            type_display = OPERATION_TYPES.get(op_type, op_type)
            ticker = op.get("ticker", op.get("figi", ""))
            name = op.get("name", op.get("description", ""))
            payment = money_value(op.get("payment"))
            commission = money_value(op.get("commission"))
            quantity = int(op.get("quantity", 0) or 0)
            price = money_value(op.get("price"))
            currency = op.get("currency", "rub").upper()
            date = parse_ts(op.get("date"))

            op_dict = {
                "date": date,
                "account_name": acc_name,
                "ticker": ticker,
                "name": name,
                "type": op_type,
                "type_display": type_display,
                "quantity": quantity if quantity else "",
                "price": price if price else "",
                "payment": payment,
                "currency": currency,
                "commission": commission,

            }
            all_operations.append(op_dict)

            # Build cashflows for XIRR
            if date:
                if op_type in DEPOSIT_TYPES:
                    account_cashflows[acc_name].append((date, -abs(payment)))  # deposit = negative
                elif op_type in WITHDRAWAL_TYPES:
                    account_cashflows[acc_name].append((date, abs(payment)))  # withdrawal = positive

    # Sort operations once — all consumers expect chronological order
    all_operations.sort(key=lambda x: x.get("date") or datetime.min)

    # ── Step 3: Calculate analytics ──
    print("\n── Расчёт аналитики ──")

    # XIRR
    print("  Считаю XIRR...")
    today = datetime.now().replace(tzinfo=None)
    total_value = sum(p.get("total_value", 0) for p in all_positions)

    # Total XIRR
    all_cashflows = []
    for acc_cfs in account_cashflows.values():
        all_cashflows.extend(acc_cfs)
    all_cashflows.append((today, total_value))
    total_xirr = xirr(all_cashflows)
    if total_xirr:
        print(f"  XIRR общий: {total_xirr:.2%}")

    # Per-account XIRR
    acc_values = defaultdict(float)
    for p in all_positions:
        acc_values[p.get("account_name", "")] += p.get("total_value", 0)

    account_xirrs = {}
    for acc_name, cfs in account_cashflows.items():
        acc_cfs = cfs + [(today, acc_values.get(acc_name, 0))]
        account_xirrs[acc_name] = xirr(acc_cfs)

    # Realized P&L
    print("  Считаю реализованный P&L (FIFO)...")
    closed_lots = calculate_realized_pnl(all_operations)
    print(f"  Закрытых лотов: {len(closed_lots)}")

    # Dividends by ticker
    divs_by_ticker = aggregate_dividends_by_ticker(all_operations)

    # Total deposits/withdrawals
    total_deposits = sum(-cf[1] for cfs in account_cashflows.values()
                         for cf in cfs if cf[1] < 0)
    total_withdrawals = sum(cf[1] for cfs in account_cashflows.values()
                            for cf in cfs if cf[1] > 0)

    # ── Step 4: Fetch candles for position sheets ──
    print("\n── Загружаю исторические цены ──")
    candles_cache = {}  # ticker -> [{date, close}, ...]

    # Collect all unique tickers (current + from operations)
    all_tickers = set()
    ticker_to_figi = {}
    for p in all_positions:
        t = p.get("ticker", "")
        if t:
            all_tickers.add(t)
            if p.get("figi"):
                ticker_to_figi[t] = p["figi"]

    for op in all_operations:
        t = op.get("ticker", "")
        if t and op.get("type") in BUY_SELL_TYPES:
            all_tickers.add(t)

    # Also get figi from instruments cache
    for uid, inst in instruments_cache.items():
        t = inst.get("ticker", "")
        f = inst.get("figi", "")
        if t and f:
            ticker_to_figi[t] = f

    # Fetch candles — parallel with ThreadPoolExecutor
    tickers_with_figi = [(t, ticker_to_figi[t]) for t in sorted(all_tickers) if t in ticker_to_figi]
    print(f"  Загружаю свечи для {len(tickers_with_figi)} инструментов (параллельно)...")

    def fetch_ticker_candles(ticker_figi):
        ticker, figi = ticker_figi
        raw_candles = get_candles(figi, ANALYSIS_START, datetime.now(timezone.utc))
        dates = []
        prices = []
        for c in raw_candles:
            d = parse_ts(c.get("time"))
            close = money_value(c.get("close"))
            if d and close:
                dates.append(d)
                prices.append(close)
        return ticker, {"dates": dates, "prices": prices}

    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {executor.submit(fetch_ticker_candles, tf): tf[0] for tf in tickers_with_figi}
        for future in as_completed(futures):
            ticker, data = future.result()
            candles_cache[ticker] = data
            print(f"  {ticker}: {len(data['dates'])} свечей")

    # ── Step 5: Reconstruct portfolio history ──
    print("\n  Реконструкция истории портфеля...")
    portfolio_history = reconstruct_portfolio_history(all_operations, candles_cache)
    print(f"  Снимков: {len(portfolio_history)}")

    # ── Step 6: Build Excel ──
    print(f"\n{'═' * 60}")
    print(f"  Всего позиций: {len(all_positions)}")
    print(f"  Всего операций: {len(all_operations)}")
    print(f"  Создаю Excel-отчёт...")
    print(f"{'═' * 60}")

    wb = Workbook()
    wb.remove(wb.active)

    # Tier 1: Dashboard
    total_realized_pnl = sum(lot["pnl"] for lot in closed_lots)
    build_dashboard(wb, all_positions, all_operations, portfolio_history,
                    total_xirr, total_deposits, total_withdrawals, divs_by_ticker,
                    total_realized_pnl)
    build_returns_sheet(wb, portfolio_history, all_operations, total_xirr, account_xirrs)
    build_cashflows_sheet(wb, all_operations, portfolio_history)

    # Tier 2: Analytics
    build_portfolio_sheet(wb, all_positions, divs_by_ticker)
    build_closed_positions_sheet(wb, closed_lots)
    build_dividends_sheet(wb, all_operations)
    build_taxes_sheet(wb, all_operations, all_positions)
    build_accounts_sheet(wb, all_positions, all_operations, account_xirrs)

    # Tier 3: Data
    build_operations_sheet(wb, all_operations)
    build_commissions_sheet(wb, all_operations)
    build_instruments_sheet(wb, instruments_cache)
    build_history_sheet(wb, portfolio_history)

    # Position sheets
    print("  Создаю листы позиций...")
    position_info_by_ticker = {}
    for p in all_positions:
        t = p.get("ticker", "")
        if t:
            position_info_by_ticker[t] = p

    ops_by_ticker = defaultdict(list)
    for op in all_operations:
        t = op.get("ticker", "")
        if t:
            ops_by_ticker[t].append(op)

    for ticker in sorted(all_tickers):
        if ticker not in candles_cache or not candles_cache[ticker].get("dates"):
            continue
        pos_info = position_info_by_ticker.get(ticker, {
            "ticker": ticker, "name": "", "sector": "", "instrument_type": "",
            "cur_price": 0, "avg_price": 0, "quantity": 0, "total_value": 0,
            "pnl": 0, "pnl_pct": 0,
        })
        build_position_sheet(wb, ticker, pos_info, ops_by_ticker.get(ticker, []),
                             candles_cache[ticker])

    # Set dashboard as active
    wb.active = 0

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"/opt/t-invest/t-invest-analytics_{timestamp}.xlsx"
    wb.save(output_path)
    print(f"\n✅ Готово! Файл сохранён: {output_path}")
    print(f"   Листов: {len(wb.sheetnames)}")
    print(f"   Размер: {os.path.getsize(output_path) / 1024:.0f} KB")


if __name__ == "__main__":
    main()
